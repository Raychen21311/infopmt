# -*- coding: utf-8 -*-
"""
sys.py â€” RFP/å¥‘ç´„ å¯©æŸ¥ï¼ˆè³‡è¨Šè™•æª¢æ ¸ç‰ˆï¼‰+ å»ºè­°å›è¦†å…§å®¹ï¼ˆLLMç‰ˆï¼‰+ æœ¬åœ°çŸ¥è­˜åº«
æ­¤ç‰ˆé‡é»ï¼ˆ2025-10-13ï¼‰ï¼š
- æª¢æ ¸æ¨¡å¼ï¼šåƒ…ä¿ç•™ã€Œä¸€æ¬¡æ€§å¯©æŸ¥ã€ã€‚
- é¡¯ç¤ºï¼šåªä¿ç•™ã€Œå·®ç•°å°ç…§è¡¨ã€èˆ‡ã€Œå»ºè­°å›è¦†å…§å®¹ï¼ˆLLMç”Ÿæˆï¼‰ã€ã€‚
- å»ºè­°å›è¦†å…§å®¹ï¼ˆLLMç”Ÿæˆï¼‰ï¼š
  * ä»¥ Prompt è¦æ±‚ LLM ç”¢å‡ºæ­£å¼ã€ç°¡æ½”çš„ã€Œå»ºè­°å›è¦†å…§å®¹ã€ã€‚
  * ç¬¬ä¸€é»å›ºå®šæ¶µè“‹ã€æœ¬æ¡ˆæ¡è³¼é‡‘é¡ï¼ˆè¬å…ƒï¼‰ã€ï¼›ç¬¬äºŒé»å›ºå®šæ¶µè“‹ã€ç¶­é‹è²»ç”¨é€å¹´éæ¸›ã€å¥ã€‚
  * å¯é¸æ“‡åŠ å…¥ã€çŸ¥è­˜åº«é …ç›®ã€ä½œç‚º Prompt çš„ä¸Šä¸‹æ–‡ï¼ˆä¸ç›´æ¥æ’å…¥åˆ°è‰ç¨¿ï¼‰ã€‚
  * é ç®—é‡‘é¡ç”± LLM å¾ RFP/å¥‘ç´„å…¨æ–‡æŠ½å–ï¼ˆè½‰ç‚ºã€è¬å…ƒã€ï¼‰ï¼Œå¯æ‰‹å‹•è¦†è“‹ã€‚
- æœ¬åœ°çŸ¥è­˜åº«ï¼š
  * å…§å»ºé è¨­é …ç›®ï¼ˆé†«ç™‚æ¨™æº–åƒè€ƒã€ç¶­é‹éæ¸›åŸå‰‡ã€å¥å‹ç¯„æœ¬ï¼‰ã€‚
  * å¯æ–°å¢ã€åˆªé™¤ã€å‹¾é¸æ˜¯å¦ç´å…¥ Promptï¼›å¯ä¸Šå‚³æ­·å²å›è¦†æ–‡å­—/PDFä½œç‚ºçŸ¥è­˜åº«é …ç›®ã€‚
- Excelï¼šåƒ…è¼¸å‡ºã€Œå·®ç•°å°ç…§ã€å·¥ä½œè¡¨ã€‚
- ä¸ä¸‹è¼‰ Wordï¼›ä¸å–å¾—æ¡ˆä»¶æ€§è³ªã€‚
"""

import os
import re
import json
import io
from typing import List, Dict, Any, Tuple

import streamlit as st
import fitz  # PyMuPDF
import pandas as pd
from difflib import SequenceMatcher
from dotenv import load_dotenv

# å¯é¸ï¼šGoogle Generative AIï¼ˆGeminiï¼‰
try:
    import google.generativeai as genai
except Exception:
    genai = None

# -------------------- LLM åˆå§‹åŒ– --------------------
load_dotenv()
API_KEY = os.getenv('GOOGLE_API_KEY')
MODEL_NAME = "gemini-2.5-flash"
model = None
if genai and API_KEY:
    try:
        genai.configure(api_key=API_KEY)
        model = genai.GenerativeModel(MODEL_NAME)
    except Exception:
        model = None

# -------------------- å¸¸æ•¸ --------------------
KB_PATH = 'kb_store.json'

# -------------------- æª”æ¡ˆå‹æ…‹ --------------------
def is_pdf(name: str) -> bool:
    return name.lower().endswith(".pdf")

def is_text(name: str) -> bool:
    return any([name.lower().endswith(ext) for ext in ('.txt', '.md', '.json')])

# ==================== æª¢æ ¸æ¸…å–®ï¼ˆå« F å…¶ä»–é‡é»ï¼‰ ====================
def build_rfp_checklist() -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []

    def add(cat, code, text):
        items.append({"category": cat, "id": code, "item": text})

    # A åŸºæœ¬èˆ‡å‰æ¡ˆ
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A0", "æœ¬æ¡ˆå±¬é–‹ç™¼å»ºç½®ã€ç³»çµ±ç¶­é‹ã€åŠŸèƒ½å¢ä¿®ã€å¥—è£è»Ÿé«”ã€ç¡¬é«”ã€å…¶ä»–?")
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A1", "æœ¬æ¡ˆç‚ºå»¶çºŒæ€§åˆç´„ï¼Œå‰æ¡ˆæ¡è³¼ç°½é™³å½±æœ¬å·²é™„ã€‚")
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A2.1", "æœ¬æ¡ˆäº‹å‰æ›¾èˆ‡è³‡è¨Šè™•è¨è«–ï¼šæœ¬æ¡ˆç›¸é—œæŠ€è¡“æ–‡ä»¶ç”±è³‡è¨Šè™•å”åŠ©æ’°å¯«ã€‚")
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A2.2", "æœ¬æ¡ˆäº‹å‰æ›¾èˆ‡è³‡è¨Šè™•è¨è«–ï¼šè¦åŠƒéšæ®µæ›¾èˆ‡è³‡è¨Šè™•é–‹æœƒè¨è«–æ¡è³¼å…§å®¹ä¸¦æœ‰æœƒè­°ç´€éŒ„ã€‚")
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A2.3", "æœ¬æ¡ˆç°½è¾¦ä¹‹å‰ï¼Œå·²ä»¥è«‹è¾¦å–®éäº¤å¥‘ç´„æ›¸ã€éœ€æ±‚èªªæ˜æ›¸ç­‰ç›¸é—œæ–‡ä»¶ï¼Œé€è«‹è³‡è¨Šè™•æª¢è¦–ï¼Œä¸¦ä¿ç•™è‡³å°‘5å€‹å·¥ä½œå¤©ä¹‹å¯©é–±æœŸå¾Œå–å¾—å›è¦†ã€‚")
    add("A åŸºæœ¬èˆ‡å‰æ¡ˆ", "A2.4", "æœ¬æ¡ˆäº‹å‰æœªèˆ‡è³‡è¨Šè™•è¨è«–ï¼ˆç„¡ï¼‰ã€‚")

    # B ç¾æ³èªªæ˜
    add("B ç¾æ³èªªæ˜", "B1.1", "æä¾›æœ€æ–°ç‰ˆç¡¬é«”è¨­å‚™åŠç¶²è·¯ä¹‹æ¶æ§‹åœ–(ä¸å«IP Address)ï¼šæ˜ç¢ºè¡¨é”ç¡¬é«”æ”¾ç½®å€åŸŸï¼ˆå«æ©Ÿæˆ¿/å€åŸŸï¼‰ã€‚")
    add("B ç¾æ³èªªæ˜", "B1.2", "æä¾›ç¶²è·¯ä»‹æ¥æ–¹å¼èˆ‡é–‹ç™¼å·¥å…·ä¹‹å» ç‰Œã€å‹è™Ÿã€ç‰ˆæœ¬ç­‰è³‡è¨Šã€‚")
    add("B ç¾æ³èªªæ˜", "B2", "ç½®æ–¼æœ¬éƒ¨æ©Ÿæˆ¿ä¹‹ç³»çµ±ï¼Œå¦‚å¦è¨­å°å¤–é€£ç·šç·šè·¯è€…ï¼Œæä¾›é€£ç·šå°è±¡ã€ç¨®é¡åŠè¦æ ¼æ¸…å–®ã€‚")
    add("B ç¾æ³èªªæ˜", "B3", "æä¾›ä½¿ç”¨è€…æˆ–ä½¿ç”¨æ©Ÿé—œä¹‹ç¤ºæ„åœ–æˆ–èªªæ˜ã€‚")
    add("B ç¾æ³èªªæ˜", "B4", "æä¾›æœ€æ–°ç¶²ç«™ç¶²å€ã€‚")
    add("B ç¾æ³èªªæ˜", "B5", "æä¾›æ‡‰ç”¨ç³»çµ±åŠŸèƒ½æ¸…å–®æˆ–æ¶æ§‹åœ–ï¼ˆå« OSã€DB åç¨±èˆ‡ç‰ˆæœ¬ï¼‰ã€‚")

    # C è³‡å®‰éœ€æ±‚
    add("C è³‡å®‰éœ€æ±‚", "C1.1", "ç¬¦åˆæœ¬éƒ¨æ¡è³¼å¥‘ç´„è¦ç¯„ä¹‹è³‡è¨Šå®‰å…¨èˆ‡å€‹è³‡ä¿è­·è¦æ±‚ï¼šå·²å¡«åˆ—å®‰å…¨ç­‰ç´šä¸”èˆ‡æœ€æ–°æ ¸å®šç­‰ç´šç›¸ç¬¦ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C1.2", "è¦æ±‚ç³»çµ±ç¬¦åˆã€Šè³‡é€šå®‰å…¨è²¬ä»»ç­‰ç´šåˆ†ç´šè¾¦æ³•ã€‹ä¹‹ã€è³‡é€šç³»çµ±é˜²è­·åŸºæº–ã€ã€SSDLC å„éšæ®µå®‰å…¨å·¥ä½œï¼›è¦æ±‚å» å•†æäº¤ã€è³‡é€šç³»çµ±é˜²è­·åŸºæº–è‡ªè©•è¡¨ã€ä¸¦å¢åˆ—ç½°å‰‡ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C1.3", "è¦æ±‚å» å•†ä¹‹æœå‹™æ°´æº–æ»¿è¶³ç³»çµ±æœ€å¤§å¯å®¹å¿ä¸­æ–·æ™‚é–“ï¼ˆRTOï¼‰ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C1.4", "éç½®æ–¼æœ¬éƒ¨æ©Ÿæˆ¿ä¹‹æ ¸å¿ƒè³‡é€šç³»çµ±ï¼Œç´å…¥ SOC ç¯„åœã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C1.5", "å§”å¤–éœ€æ±‚æ¶‰åŠè³‡é€šæŠ€è¡“æœå‹™ï¼ˆå¦‚é›²ç«¯ï¼‰å·²è©•ä¼°åˆæ³•æ€§ã€æŠ€è¡“ç¶­é‹ã€æ³•éµèˆ‡æ¬Šåˆ©ç¾©å‹™æ­¸å±¬ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C1.6", "è¦æ±‚å» å•†ä¸å¾—ä½¿ç”¨æˆ–è¨­è¨ˆä¸ç¬¦å®‰å…¨è¦ç¯„ä¹‹å¸³è™Ÿå¯†ç¢¼ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C2.1", "å·¨é¡/è³‡å®‰æ¡è³¼æˆ–é«˜ç´šå®‰å…¨ç­‰ç´šæ¡ˆä»¶ï¼šæŠ•æ¨™å» å•†å…·å‚™å®‰å…¨è»Ÿé«”é–‹ç™¼èƒ½åŠ›ä¸¦é€šéè³‡å®‰ç®¡ç†ç³»çµ±é©—è­‰ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C2.2", "å·¨é¡/è³‡å®‰/é«˜ç´šï¼šå°ˆæ¡ˆç®¡ç†äººå“¡è‡³å°‘1äººå…·è³‡è¨Šå®‰å…¨å°ˆæ¥­èªè­‰ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C2.3", "å·¨é¡/è³‡å®‰/é«˜ç´šï¼šå°ˆæ¡ˆæŠ€è¡“äººå“¡è‡³å°‘1äººå…·ç¶²è·¯å®‰å…¨æŠ€èƒ½ä¹‹è¨“ç·´è­‰æ›¸æˆ–è­‰ç…§ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C3.1", "å…è¨±åˆ†åŒ…è€…ï¼šåˆ†åŒ…å» å•†é ˆæ¯”ç…§æ‰¿åŒ…å» å•†å…±åŒéµå®ˆè³‡å®‰è¦å®šã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C3.2", "å…è¨±åˆ†åŒ…è€…ï¼šæŠ•æ¨™å» å•†æ–¼æœå‹™å»ºè­°æ›¸æ•˜æ˜åˆ†åŒ…å» å•†åŸºæœ¬è³‡æ–™ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C4", "ä¸å¾—æ¡ç”¨å¤§é™¸å» ç‰Œè³‡é€šè¨Šç”¢å“ï¼ˆå¥‘ç´„è‰æ¡ˆç¬¬å…«æ¢(å…­)åŠ(äºŒäº”)ï¼‰ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C5", "ç¬¦åˆã€è³‡é€šç³»çµ±ç±Œç²å„éšæ®µè³‡å®‰å¼·åŒ–æªæ–½åŸ·è¡Œæª¢æ ¸è¡¨ã€ï¼ˆé–‹ç™¼é™„è¡¨1/ç¶­é‹é™„è¡¨2ï¼‰ã€‚")
    add("C è³‡å®‰éœ€æ±‚", "C6", "è³‡æ–™åº«ä¸­æ©Ÿæ•è³‡æ–™å·²æ¡ç”¨æˆ–è¦åŠƒé©ç•¶åŠ å¯†æŠ€è¡“ã€‚")

    # D ä½œæ¥­éœ€æ±‚ï¼ˆç¯€éŒ„ï¼‰
    add("D ä½œæ¥­éœ€æ±‚", "D1", "åˆ—å‡ºæ‰€éœ€è»Ÿç¡¬é«”èˆ‡ç¶²è·¯è¨­å‚™æ¸…å–®ï¼Œèªªæ˜ä½¿ç”¨è³‡è¨Šè™•è¨­å‚™/æ—¢æœ‰è¨­å‚™æˆ–å¦è¡Œæ¡è³¼ï¼ˆå„ªå…ˆ VM/å…±åŒä¾›æ‡‰å¥‘ç´„ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D2", "ç³»çµ±é–‹ç™¼æˆ–åŠŸèƒ½å¢ä¿®æ‡‰åˆ—å‡ºæ‰€éœ€ç³»çµ±åŠŸèƒ½ï¼ˆåœ°æ–¹æ”¿åºœç³»çµ±å»ºè­°æä¾›è³‡æ–™ä¸‹è¼‰æˆ–ä»‹æ¥ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D3", "æ•˜æ˜è³‡è¨Šç³»çµ±èˆ‡å…¶ä»–è»Ÿé«”ç³»çµ±ä¹‹ç›¸äº’é—œä¿‚ä¸¦èªªæ˜æ‰€æœ‰åˆ©å®³é—œä¿‚äººã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D4", "æä¾›æ°‘çœ¾ä¸‹è¼‰æª”æ¡ˆè€…ï¼Œå¢åŠ  ODF æ ¼å¼ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D5", "é–‹ç™¼ App å·²é–±è®€ä¸¦éµå¾ªåœ‹ç™¼æœƒç›¸é—œè¦å®šï¼ˆé™„ä»¶2ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D6", "é–‹ç™¼ App ç¬¦åˆé€šå‚³æœƒã€App ç„¡éšœç¤™é–‹ç™¼æŒ‡å¼•ã€ä¸¦å¡«å ±æª¢æ ¸è¡¨ï¼ˆé™„ä»¶3ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D7", "ç¶²ç«™æœå‹™ä¹‹ç³»çµ±ç¬¦åˆåœ‹ç™¼æœƒã€æ”¿åºœç¶²ç«™æœå‹™ç®¡ç†è¦ç¯„ã€ä¸¦å¡«å ±æª¢æ ¸è¡¨ï¼ˆé™„ä»¶4ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D8", "é‡å°æ¥­å‹™æˆ–å€‹äººè³‡æ–™ï¼Œæä¾›å¾ŒçºŒ OpenData æˆ– MyData æœå‹™å»ºè­°ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D9", "ç³»çµ±ç¶­è­·åŒ…å«å®šæœŸåˆ°å ´ã€ç·Šæ€¥åˆ°å ´ã€è«®è©¢æœå‹™ï¼›SLA èˆ‡ç¸¾æ•ˆæŒ‡æ¨™é€£å‹•ä¸¦è¨­è¨ˆä½¿ç”¨è€…æ»¿æ„åº¦èª¿æŸ¥ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D10", "å±¥ç´„æœå‹™éŠœæ¥å¥‘ç´„æœŸé–“ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D11", "é–‹ç™¼åŠæ¸¬è©¦è¨­å‚™èˆ‡ç’°å¢ƒéœ€æ±‚èªªæ˜ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D12", "æ•™è‚²è¨“ç·´åŠå®¢æœæœå‹™ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D13", "ä¿å›ºæœå‹™ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D14", "ç”¢å“æˆæ¬Š (License) ç¬¦åˆéœ€æ±‚ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D15", "ä½œæ¥­éœ€æ±‚å¿…é ˆç´å…¥ä¹‹åˆ¶å¼æ–‡å¥ï¼ˆè©³è¨»5ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D16", "å¦‚æœ‰ GIS / OpenData / MyData ä½œæ¥­éœ€æ±‚ï¼Œç´å…¥ä¹‹åˆ¶å¼æ–‡å¥ï¼ˆè©³è¨»6ï¼‰ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D17", "ä¸Šç·šå‰å®Œæˆéœ€æ±‚è¨ªè«‡ã€éœ€æ±‚ç¢ºèªèˆ‡æ¸¬è©¦ï¼ˆå«æ•ˆèƒ½æ¸¬è©¦ï¼‰ï¼›æäº¤æ¸¬è©¦è¨ˆç•«èˆ‡æ¸¬è©¦å ±å‘Šã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D18", "æ¶‰åŠé†«ç™‚/å¥åº·è³‡æ–™äº¤æ›è€…ï¼Œç´å…¥ FHIR äº¤æ›æ¨™æº–ã€‚")
    add("D ä½œæ¥­éœ€æ±‚", "D19", "åŠŸèƒ½éœ€æ±‚è¨­è¨ˆè€ƒé‡å°å…¥ AI ä»¥ç¯€çœäººåŠ›/é¿å…éŒ¯èª¤èˆ‡æ±ºç­–åˆ†æåŠé¢¨éšªé è­¦ã€‚")

    # E ç”¢å“äº¤ä»˜
    add("E ç”¢å“äº¤ä»˜", "E1", "äº¤ä»˜æ™‚ç¨‹åˆç†ï¼Œä¸¦èˆ‡é–‹ç™¼æ–¹å¼ï¼ˆç€‘å¸ƒ/æ•æ·ï¼‰ä¸€è‡´ã€‚")
    add("E ç”¢å“äº¤ä»˜", "E2", "é–‹ç™¼/å¢ä¿®äº¤ä»˜å“å®Œæ•´ï¼ˆå°ˆæ¡ˆè¨ˆç•«ã€éœ€æ±‚/è¨­è¨ˆã€æ¸¬è©¦è¨ˆç•«/å ±å‘Šã€å»ºç½®è¨ˆç•«ã€æ‰‹å†Šã€æ•™è‚²è¨“ç·´ã€ä¿å›ºç´€éŒ„ã€åŸå§‹ç¢¼/åŸ·è¡Œç¢¼ã€æœ€é«˜æ¬Šé™å¸³å¯†ã€è‡ªè©•è¡¨èˆ‡é›»å­æª”ï¼‰ã€‚")
    add("E ç”¢å“äº¤ä»˜", "E3", "ç¶­è­·äº¤ä»˜å“ï¼ˆå°ˆæ¡ˆåŸ·è¡Œè¨ˆç•«ã€ç¶­è­·å·¥ä½œå ±å‘Šã€æœ€æ–°ç‰ˆè¨­è¨ˆ/æ‰‹å†Šã€æœ€æ–°ç‰ˆåŸå§‹ç¢¼/åŸ·è¡Œç¢¼ã€è‡ªè©•è¡¨èˆ‡é›»å­æª”ï¼‰ã€‚")
    add("E ç”¢å“äº¤ä»˜", "E4", "å¿…é ˆç´å…¥ä¹‹åˆ¶å¼æ–‡å¥ï¼ˆè©³è¨»8ï¼‰ï¼šäº¤ä»˜ä¹‹åŸå§‹ç¨‹å¼ç¢¼ã€åŸ·è¡Œç¢¼ï¼Œæœ¬éƒ¨å¾—è¦æ±‚æ‰¿åŒ…å» å•†æ–¼æœ¬éƒ¨æŒ‡å®šä¹‹ç’°å¢ƒé€²è¡Œå†ç”Ÿæ¸¬è©¦ï¼Œä¸¦æ‡‰æä¾›æ‰€ä½¿ç”¨ä¹‹é–‹ç™¼å·¥å…·ï¼Œä»¥é©—è­‰å…¶æ­£ç¢ºæ€§ã€‚")
    add("E ç”¢å“äº¤ä»˜", "E5", "ç¶²è·¯è¨­å‚™è³¼ç½®æ™‚ï¼Œé©—æ”¶ä»¥å½Œå°æ–¹å¼äº¤ä»˜å¸³å¯†ã€è¨­å®šæª”ã€è¦å‰‡åˆ—è¡¨èˆ‡æ¶æ§‹ç­‰ã€‚")

    # F å…¶ä»–é‡é»ï¼ˆä¾é å¯©è¡¨ï¼‰
    add("F å…¶ä»–é‡é»", "F1", "ä½¿ç”¨æœ¬éƒ¨æ©Ÿæˆ¿ VMã€å…±ç”¨è³‡æ–™åº«ï¼Œå·²å®Œæˆè©•ä¼°åŠæˆæœ¬åˆ†æ”¤è¡¨å¡«å¯«ï¼Œä¸¦å·²åˆ†æ”¤ç¶“è²»ã€‚")
    add("F å…¶ä»–é‡é»", "F2", "ç¶“è²»é ä¼°ä¹‹åˆç†æ€§åŠç¶“è³‡é–€æ­¸é¡ä¹‹æ­£ç¢ºæ€§ï¼›ç¶­è­·è²»ç”¨è¨ˆç®—æ¯”ç‡æ‡‰é€å¹´éæ¸›ï¼ˆä¸å«æ—¢æœ‰æ“´å¢ä¸”éä¿å›ºéƒ¨åˆ†ï¼‰ã€‚")
    add("F å…¶ä»–é‡é»", "F3", "æ¡è³¼å…§å®¹ç„¡å‰å¾Œä¸ä¸€è‡´æƒ…å½¢ã€‚")
    add("F å…¶ä»–é‡é»", "F4", "å°ç…§ä½œæ¥­éœ€æ±‚æª¢æŸ¥å¥‘ç´„æ›¸ã€Œæœå‹™æ°´æº–åŠç¸¾æ•ˆé•ç´„é‡‘ã€ä¹‹å…§å®¹æœ‰ç„¡ç¼ºæ¼ã€‚")
    add("F å…¶ä»–é‡é»", "F5", "æ¡è³¼å¥‘ç´„æ›¸ã€Œå±¥ç´„æ¨™çš„ã€å…§å®¹æ­£ç¢ºï¼Œç„¡ç¼ºæ¼ã€‚")
    add("F å…¶ä»–é‡é»", "F6", "é–‹ç™¼æˆ–å¢ä¿®ç³»çµ±ä¹‹ä»‹æ¥å…§å®¹ï¼Œå·²æ´½ç›¸é—œå–®ä½åŒæ„ï¼Œä¸¦ç¢ºèªå°æ–¹ç³»çµ±å¢ä¿®åŠç¶“è²»ä¾†æºã€‚")
    add("F å…¶ä»–é‡é»", "F7", "ç›®å‰ä½¿ç”¨ä¹‹ç¡¬é«”è¨­å‚™æ–¼å±¥ç´„å®Œæˆå¾Œï¼Œå¦‚æ±°æ›æˆ–ä¸å†ä½¿ç”¨è€…ï¼Œè¦åŠƒä¸‹æ¶æ—¥æœŸã€‚")
    add("F å…¶ä»–é‡é»", "F8", "æº–ç”¨æœ€æœ‰åˆ©æ¨™ä¹‹è©•é¸é …ç›®èˆ‡é…åˆ†ï¼Œé™„éŒ„å¯ä¾éœ€æ±‚èª¿æ•´æ•´ä½µä¸¦åŒæ­¥ä¿®æ­£ç›¸é—œé™„éŒ„ã€‚")

    return items

# ==================== PDF è§£æ ====================
def extract_text_with_headers(pdf_bytes: bytes, filename: str) -> str:
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    parts = []
    for i, page in enumerate(doc, start=1):
        text = page.get_text('text').strip()
        if not text:
            blocks = page.get_text('blocks')
            text = "\n\n".join([b[4].strip() for b in blocks if b[4].strip()])
        parts.append(f"\n\n===== ã€æª”æ¡ˆ: {filename} é : {i}ã€‘ =====\n" + text)
    return "\n".join(parts)

# ==================== LLM Promptsï¼ˆæª¢æ ¸/é å¯©è§£æï¼‰ ====================
def make_batch_prompt(batch_code: str, items: List[Dict[str, Any]], corpus_text: str) -> str:
    checklist_lines = "\n".join([f"{it['id']}ï½œ{it['item']}" for it in items])
    return f"""
ä½ æ˜¯æ”¿åºœæ©Ÿé—œè³‡è¨Šè™•ä¹‹æ¡è³¼/RFP/å¥‘ç´„å¯©æŸ¥å§”å“¡ã€‚è«‹ä¾ä¸‹åˆ—ã€Œæª¢æ ¸æ¢ç›®ï¼ˆ{batch_code} æ‰¹ï¼‰ã€é€æ¢å¯©æŸ¥æ–‡ä»¶å…§å®¹ä¸¦å›å‚³**å”¯ä¸€ JSON é™£åˆ—**ï¼Œé™£åˆ—å…§æ¯å€‹å…ƒç´ å°æ‡‰ä¸€æ¢æ¢ç›®ã€‚
ã€å¯©æŸ¥åŸå‰‡ã€‘
1) åƒ…ä¾æ–‡ä»¶æ˜è¼‰å…§å®¹åˆ¤æ–·ï¼›æœªæåŠå³æ¨™ç¤ºã€ŒæœªæåŠã€ã€‚
2) è‹¥å±¬ä¸é©ç”¨ï¼ˆä¾‹ï¼šæœªå…è¨±åˆ†åŒ…ï¼‰ï¼Œè«‹å›ã€Œä¸é©ç”¨ã€ä¸¦èªªæ˜ä¾æ“šã€‚
3) å‹™å¿…å¼•ç”¨åŸæ–‡çŸ­å¥èˆ‡æª”å/é ç¢¼ä½œç‚º evidenceã€‚
4) ***åš´ç¦è¼¸å‡ºä»»ä½•èˆ‡è¦æ ¼è¯çµ¡äººã€é›»è©±ã€å§“åã€è¯ç¹«æ–¹å¼æœ‰é—œçš„æ–‡å­—ï¼Œå³ä½¿åŸå§‹æ–‡ä»¶å…§æœ‰ã€‚***
ã€è¼¸å‡ºæ ¼å¼ â€” åƒ…èƒ½è¼¸å‡º JSON é™£åˆ—ã€‘
[
  {{
    "id": "A1",
    "category": "A åŸºæœ¬èˆ‡å‰æ¡ˆ",
    "item": "æ¢ç›®åŸæ–‡ï¼ˆå®Œæ•´è¤‡è£½ï¼‰",
    "compliance": "è‹¥ id = 'A0'ï¼šå…­é¸ä¸€ã€é–‹ç™¼å»ºç½®ï½œç³»çµ±ç¶­é‹ï½œåŠŸèƒ½å¢ä¿®ï½œå¥—è£è»Ÿé«”ï½œç¡¬é«”ï½œå…¶ä»–ã€‘ï¼›è‹¥ id â‰  'A0'ï¼šå››é¸ä¸€ã€ç¬¦åˆï½œéƒ¨åˆ†ç¬¦åˆï½œæœªæåŠï½œä¸é©ç”¨ã€‘",
    "evidence": {{ "file": "æª”å", "page": "é ç¢¼", "quote": "é€å­—å¼•è¿°" }},
    "recommendation": "è‹¥æœªæåŠ/éƒ¨åˆ†ç¬¦åˆï¼Œè«‹çµ¦å…·é«”è£œå¼·æ–¹å‘ï¼›å¦å‰‡ç•™ç©º"
  }}
]
ã€æœ¬æ‰¹æª¢æ ¸æ¸…å–®ï¼ˆidï½œitemï¼‰ã€‘
{checklist_lines}
ã€æ–‡ä»¶å…¨æ–‡ï¼ˆå«æª”å/é ç¢¼æ¨™è¨»ï¼‰ã€‘
{corpus_text}
""".strip()


def make_precheck_parse_prompt(corpus_text: str) -> str:
    return f"""
ä½ æ˜¯æ”¿åºœæ©Ÿé—œè³‡è¨Šè™•ä¹‹æ¡è³¼å¯©æŸ¥åŠ©ç†ã€‚ä»¥ä¸‹æ˜¯ä¸€ä»½æˆ–å¤šä»½ã€ŒåŸ·è¡Œå–®ä½é å…ˆå¯©æŸ¥è¡¨ã€çš„ PDF æ–‡å­—ï¼ˆå·²æ¨™è¨»æª”åèˆ‡é ç¢¼ï¼‰ã€‚
è«‹å°‡è¡¨æ ¼/æ¢åˆ—é€åˆ—è½‰ç‚º **JSON é™£åˆ—**ï¼Œæ¯åˆ—ä¸€ç­†ï¼Œåƒ…è¼¸å‡ºä¸‹åˆ—æ¬„ä½ï¼š
- "id"ï¼ˆç²—ç·¨è™Ÿï¼Œå¯ç©ºï¼‰
- "item"ï¼ˆæª¢æ ¸é …ç›®ï¼‰
- "status"ï¼ˆåƒ…èƒ½ã€ç¬¦åˆï½œä¸é©ç”¨ã€‘æˆ–ç©ºå­—ä¸²ï¼‰
- "biz_ref_note"ï¼ˆå°æ‡‰é æ¬¡/å‚™è¨»ï¼‰
- "std_id"ï¼ˆè‹¥èƒ½åˆ¤å®šå°æ‡‰æ¸…å–®ç·¨è™Ÿï¼‰
Evidence è‡³å°‘ä¸€ç­†ï¼š{{"file":"...", "page": é ç¢¼, "quote":"..."}}
ç¦æ­¢è¼¸å‡ºä»»ä½•è¯çµ¡è³‡è¨Šï¼ˆå§“åã€é›»è©±ã€Email ç­‰ï¼‰ã€‚
ã€è¼¸å‡ºæ ¼å¼ â€” åƒ…èƒ½è¼¸å‡º JSON é™£åˆ—ã€‘
[ {{ "id": "ç¾æ³èªªæ˜-1.(2)", "item": "...", "status": "ç¬¦åˆ", "biz_ref_note": "...", "std_id": "B1.2" }} ]
ã€æ–‡ä»¶å…¨æ–‡ï¼ˆå«æª”å/é ç¢¼æ¨™è¨»ï¼‰ã€‘
{corpus_text}
""".strip()

# ==================== æœ¬åœ°çŸ¥è­˜åº« ====================
def default_kb_items() -> List[Dict[str, Any]]:
    return [
        {
            "id": "kb_med_std",
            "title": "é†«ç™‚è³‡æ–™æ¨™æº–åƒè€ƒ",
            "body": (
                "æœ‰é—œé†«ç™‚è³‡æ–™å…§å®¹ï¼Œè«‹åƒé–±æœ¬éƒ¨é†«ç™‚è³‡è¨Šå¤§å¹³å°ä¹‹é†«ç™‚è³‡è¨Šæ¨™æº–ï¼Œå¦‚ FHIRã€LOINCã€SNOMED CTã€RxNormï¼Œ"
                "ä¸”ç¬¦åˆä¸‰å¤§AIä¸­å¿ƒã€SMART on FHIRç­‰ä½œæ¥­äº‹é …ã€‚å¦å¦‚æœ‰TWCDIåŠIGéœ€æ±‚ï¼Œå¯è‡³è©²å¹³å°ææ¡ˆã€‚"
            ),
            "tags": ["é†«ç™‚", "æ¨™æº–", "FHIR"],
            "default_include": False
        },
        {
            "id": "kb_maint_decrease",
            "title": "ç¶­é‹è²»ç”¨é€å¹´éæ¸›åŸå‰‡",
            "body": (
                "è³‡è¨Šç³»çµ±ä¹‹ç¶­é‹è²»ç”¨æ‡‰é€å¹´éæ¸›ï¼Œå» å•†å ±åƒ¹å¦‚æœ‰å¢é•·ï¼Œå¯è«‹å» å•†æ–¼æœ¬æ¡ˆä¹‹æœŸæœ«å ±å‘Šæä¾›ç³»çµ±ä½¿ç”¨æ•ˆç›ŠæŒ‡æ¨™ï¼Œ"
                "åšç‚ºæ¬¡å¹´ç¶­é‹è²»ç”¨æˆé•·ä¹‹åˆ¤æ–·ã€‚"
            ),
            "tags": ["ç¶­é‹", "è²»ç”¨", "é€å¹´éæ¸›"],
            "default_include": True
        },
        {
            "id": "kb_budget_sentence",
            "title": "æ¡è³¼é‡‘é¡å¥å‹ç¯„æœ¬",
            "body": "æœ¬æ¡ˆæ¡è³¼é‡‘é¡{BUDGET}è¬å…ƒï¼Œ{SUMMARY}ã€‚",
            "tags": ["é ç®—", "å¥å‹"],
            "default_include": True
        },
    ]


def load_kb() -> List[Dict[str, Any]]:
    if not os.path.exists(KB_PATH):
        return default_kb_items()
    try:
        with open(KB_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    return default_kb_items()


def save_kb(items: List[Dict[str, Any]]):
    try:
        with open(KB_PATH, 'w', encoding='utf-8') as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def kb_to_context(items: List[Dict[str, Any]], selected_ids: List[str]) -> str:
    # ä¾é¸å–é …ç›®çµ„æˆä¸Šä¸‹æ–‡æ–‡å­—ï¼ˆä¸ç›´æ¥æ’å…¥è‰ç¨¿ï¼Œåƒ…ä¾› Prompt åƒè€ƒï¼‰
    picked = []
    for it in items:
        if it.get('id') in selected_ids or it.get('default_include'):
            picked.append(f"- {it.get('title')}ï¼š{it.get('body')}")
    return "\n".join(picked)

# ==================== LLM å–å¾—é ç®—é‡‘é¡ï¼ˆè¬å…ƒï¼‰ ====================
def llm_extract_budget(corpus_text: str) -> Tuple[str, Dict[str, Any]]:
    """
    ç”± LLM å¾ RFP/å¥‘ç´„å…¨æ–‡æŠ½å–ã€æœ¬æ¡ˆæ¡è³¼/é ç®—é‡‘é¡ã€ï¼Œå›å‚³è¬å…ƒï¼ˆå­—ä¸²ï¼‰èˆ‡ evidenceã€‚
    å¤±æ•—å‰‡å› ("XXX", {}).
    """
    prompt = f"""
ä½ æ˜¯æ”¿åºœæ©Ÿé—œæ¡è³¼æ–‡ä»¶å¯©æŸ¥åŠ©ç†ã€‚è«‹ç”±ä¸‹æ–¹ã€RFP/å¥‘ç´„å…¨æ–‡ã€ä¸­å°‹æ‰¾æœ€å¯èƒ½ä»£è¡¨ã€æœ¬æ¡ˆæ¡è³¼/é ç®—é‡‘é¡ã€çš„ä¸€è™•æ•¸å­—ï¼Œ
ä¸¦å°‡å…¶æ›ç®—ç‚ºã€è¬å…ƒã€å¾Œå›å‚³å”¯ä¸€ JSON ç‰©ä»¶ï¼Œæ ¼å¼å¦‚ä¸‹ï¼Œä¸è¦è¼¸å‡ºä»»ä½•å…¶ä»–æ–‡å­—ï¼š
{{
  "budget_million": "1200", # ä¸å«å–®ä½ã€åªå¡«æ•¸å­—å­—ä¸²ï¼›è‹¥æ‰¾ä¸åˆ°è«‹å¡« "XXX"
  "evidence": {{
    "file": "æª”å",
    "page": 3,
    "quote": "é€å­—å¼•è¿°ï¼ˆä¸è¶…é200å­—ï¼‰"
  }}
}}
æ›ç®—è¦å‰‡ï¼šæ–°è‡ºå¹£/NTD/NT$/å…ƒ/è¬/ç™¾è¬/å„„ â†’ ä¸€å¾‹è½‰ç‚ºã€è¬å…ƒã€ã€‚
è‹¥åŒæ™‚å‡ºç¾ä¸åŒé‡‘é¡ï¼Œå„ªå…ˆæŒ‘é¸ä»¥ã€Œé ç®—/ç¶“è²»/æ¡è³¼é‡‘é¡ã€ç­‰é—œéµè©å°±è¿‘å‡ºç¾è€…ã€‚
ç¦æ­¢è¼¸å‡ºä»»ä½•è¯çµ¡è³‡è¨Šï¼ˆå§“åã€é›»è©±ã€Email ç­‰ï¼‰ã€‚
ã€RFP/å¥‘ç´„å…¨æ–‡ï¼ˆå«æª”å/é ç¢¼æ¨™è¨»ï¼‰ã€‘
{corpus_text}
""".strip()

    try:
        if model is None:
            return "XXX", {}
        resp = model.generate_content(prompt)
        data = json.loads((resp.text or "").strip())
        val = str(data.get("budget_million", "")).strip()
        if not val or not re.match(r"^\d+(?:\.\d+)?$", val):
            return "XXX", {}
        budget_str = str(int(round(float(val))))
        ev = data.get("evidence") or {}
        return budget_str, {"file": ev.get("file", ""), "page": ev.get("page", None), "quote": ev.get("quote", "")}
    except Exception:
        return "XXX", {}

# ==================== LLM Promptï¼šå»ºè­°å›è¦†å…§å®¹ ====================
def make_reply_prompt(corpus_text: str, kb_context: str, budget_wan: str, work_summary: str, max_points: int) -> str:
    return f"""
ä½ æ˜¯æ”¿åºœæ©Ÿé—œè³‡è¨Šè™•ä¹‹æ¡è³¼/RFP/å¥‘ç´„å¯©æŸ¥å§”å“¡ï¼Œè«‹ç”¨ç¹é«”ä¸­æ–‡æ’°å¯«ã€å»ºè­°å›è¦†å…§å®¹ã€ï¼Œé¢¨æ ¼éœ€**æ­£å¼ã€ç²¾ç°¡ã€å¯ç›´æ¥è²¼ç”¨**ï¼Œä¸¦ä»¥**ç·¨è™Ÿæ¢åˆ—**ã€‚
è«‹åš´æ ¼éµå®ˆï¼š
1) ç¬¬ä¸€é»å›ºå®šç‚ºï¼šã€Œæœ¬æ¡ˆæ¡è³¼é‡‘é¡{budget_wan}è¬å…ƒã€ï¼Œè‹¥æä¾›äº†å·¥ä½œæ‘˜è¦å‰‡è£œè¿°ï¼Œä¾‹å¦‚ã€Œå·¥ä½œå…§å®¹ç‚º{work_summary}ã€ã€‚
2) ç¬¬äºŒé»å›ºå®šç‚ºï¼šã€Œè³‡è¨Šç³»çµ±ä¹‹ç¶­é‹è²»ç”¨æ‡‰é€å¹´éæ¸›â€¦ã€ï¼ˆè¦‹ä¸‹æ–¹çŸ¥è­˜åº«å…§å®¹ï¼‰ã€‚
3) å…¶é¤˜å„é»ï¼ˆæœ€å¤š {max_points-2} é»ï¼‰è¦–æ–‡ä»¶å·®ç•°æˆ–ç¼ºæ¼ï¼Œçµ¦å‡º**å…·é«”å¯æ“ä½œ**çš„è£œå……/ä¿®æ­£å»ºè­°ï¼Œé¿å…ç©ºæ³›ã€‚
4) å…¨æ–‡ä¸å¾—è¼¸å‡ºä»»ä½•è¯çµ¡è³‡è¨Šï¼ˆå§“åã€é›»è©±ã€Email ç­‰ï¼‰ã€‚ä¸å¾—ç·¨é€ æ–‡ä»¶æœªè¼‰æ˜ä¹‹é‡‘é¡æˆ–äººåã€‚
5) åƒ…è¼¸å‡º**æ¢åˆ—æ–‡å­—**ï¼Œä¸è¦åŠ å…¥å‰è¨€ã€è½æ¬¾æˆ–è‡´æ„ã€‚
ã€å¯ä¾›åƒè€ƒä¹‹çŸ¥è­˜åº«å…§å®¹ï¼ˆåƒ…ä½œç‚ºä¸Šä¸‹æ–‡ï¼Œä¸å¿…é€å­—è²¼å…¥ï¼‰ã€‘
{kb_context}
ã€RFP/å¥‘ç´„å…¨æ–‡ï¼ˆå«æª”å/é ç¢¼æ¨™è¨»ï¼‰ã€‘
{corpus_text}
""".strip()

# ==================== è§£æ/è½‰è¡¨å·¥å…· ====================
def parse_json_array(text: str) -> List[Dict[str, Any]]:
    t = (text or "").strip()
    t = re.sub(r'^```(?:json)?', '', t, flags=re.I).strip()
    t = re.sub(r'```$', '', t, flags=re.I).strip()
    start = t.find('[')
    end = t.rfind(']')
    if start != -1 and end != -1 and end > start:
        t = t[start:end+1]
    try:
        data = json.loads(t)
        if isinstance(data, dict):
            data = [data]
        return data if isinstance(data, list) else []
    except Exception:
        return []

def normalize_status_equiv(s: str) -> str:
    if s is None:
        return "æœªæåŠ"
    t = re.sub(r"\s+", "", str(s)).lower()
    if t == "":
        return "æœªæåŠ"
    if t in ("ç¬¦åˆ", "ok", "pass", "é€šé"):
        return "ç¬¦åˆ"
    if t in ("ä¸é©ç”¨", "na", "n/a"):
        return "ä¸é©ç”¨"
    return "æœªæåŠ"

STD_ID_PATTERN = re.compile(r"^[A-F]\d+(?:\.\d+)?$")
SECTION_TO_LETTER = {"A åŸºæœ¬èˆ‡å‰æ¡ˆ": "A", "ç¾æ³èªªæ˜": "B", "è³‡å®‰éœ€æ±‚": "C", "ä½œæ¥­éœ€æ±‚": "D", "ç”¢å“äº¤ä»˜": "E", "å…¶ä»–é‡é»": "F"}
ROMAN_TO_LETTER = {"ä¸€": "A", "äºŒ": "B", "ä¸‰": "C", "å››": "D", "äº”": "E", "å…­": "F"}

def compute_std_id(raw_id: str, item: str) -> str:
    s = (raw_id or "").strip()
    if STD_ID_PATTERN.match(s):
        return s
    src = f"{raw_id} {item}".strip()
    sec = ""
    for zh, letter in SECTION_TO_LETTER.items():
        if zh in src:
            sec = letter
            break
    if not sec:
        for zh, letter in ROMAN_TO_LETTER.items():
            if f"{zh}ã€" in src or f"{zh} " in src:
                sec = letter
                break
    m1 = re.search(r"-(\d+)", raw_id or "") or re.search(r"\((\d+)\)", src)
    n1 = m1.group(1) if m1 else None
    m2 = re.search(r"\((\d+)\)", src)
    n2 = m2.group(1) if m2 else None
    if sec and n1:
        return f"{sec}{n1}" + (f".{n2}" if n2 else "")
    return ""

def parse_precheck_json(text: str) -> List[Dict[str, Any]]:
    data = parse_json_array(text)
    rows = []
    for r in data if isinstance(data, list) else []:
        if not isinstance(r, dict):
            continue
        rows.append({
            "raw_id": (r.get("id","") or "").strip(),
            "item": (r.get("item","") or "").strip(),
            "status": (r.get("status","") or "").strip(),
            "biz_ref_note": (r.get("biz_ref_note","") or "").strip(),
            "std_id": (r.get("std_id","") or "").strip(),
        })
    return rows

def precheck_rows_to_df(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    std_ids = []
    for r in rows:
        sid = r.get("std_id", "")
        if not sid:
            sid = compute_std_id(r.get("raw_id",""), r.get("item",""))
        std_ids.append(sid)
    df = pd.DataFrame({
        "ç·¨è™Ÿ": std_ids,
        "æª¢æ ¸é …ç›®": [r.get("item","") for r in rows],
        "é å¯©åˆ¤å®š": [r.get("status","") for r in rows],
        "å°æ‡‰é æ¬¡/å‚™è¨»": [r.get("biz_ref_note","") for r in rows],
    })
    df["_é å¯©ç­‰åƒ¹ç´š_éš±è—"] = df["é å¯©åˆ¤å®š"].apply(normalize_status_equiv)
    return df

# ==================== ç³»çµ±æª¢æ ¸ â†’ DataFrame ====================
def to_dataframe(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for r in results:
        rows.append({
            "é¡åˆ¥": r.get("category",""),
            "ç·¨è™Ÿ": r.get("id",""),
            "æª¢æ ¸é …ç›®": r.get("item",""),
            "ç¬¦åˆæƒ…å½¢": r.get("compliance",""),
            "ä¸»è¦è­‰æ“š": r.get("evidence", {}),
            "æ”¹å–„å»ºè­°": r.get("recommendation",""),
        })
    df = pd.DataFrame(rows)
    # ç©©å¥æ’åºï¼šA-F ä¸»ç¢¼ + æ•¸å­— + æ¬¡ç¢¼
    try:
        df["ä¸»ç¢¼"]   = df["ç·¨è™Ÿ"].str.extract(r'^([A-F])')[0]
        df["å­ç¢¼ä¸»"] = pd.to_numeric(df["ç·¨è™Ÿ"].str.extract(r'^A-F')[0], errors='coerce')
        df["å­ç¢¼æ¬¡"] = pd.to_numeric(df["ç·¨è™Ÿ"].str.extract(r'^[A-F]\d+\.(\d+)')[0], errors='coerce')
        code_order = {"A":0,"B":1,"C":2,"D":3,"E":4,"F":5}
        df["ä¸»åº"]  = df["ä¸»ç¢¼"].map(code_order).fillna(9)
        df = df.sort_values(["ä¸»åº","å­ç¢¼ä¸»","å­ç¢¼æ¬¡","ç·¨è™Ÿ"], kind='mergesort').drop(columns=["ä¸»ç¢¼","å­ç¢¼ä¸»","å­ç¢¼æ¬¡","ä¸»åº"])
    except Exception:
        pass
    return df

# ==================== å·®ç•°å°ç…§ ====================
def fuzzy_match(best_of: List[str], query: str) -> Tuple[str, float]:
    best_id, best_ratio = "", 0.0
    for cand in best_of:
        r = SequenceMatcher(a=query or "", b=cand or "").ratio()
        if r > best_ratio:
            best_ratio, best_id = r, cand
    return best_id, best_ratio

def build_compare_table(sys_df: pd.DataFrame, pre_df: pd.DataFrame) -> pd.DataFrame:
    sys_idx: Dict[str, Dict[str, Any]] = {}
    for _, row in sys_df.iterrows():
        rid = str(row.get("ç·¨è™Ÿ", "")).strip()
        if rid:
            sys_idx[rid] = row.to_dict()

    rows_out: List[Dict[str, Any]] = []
    if "_é å¯©ç­‰åƒ¹ç´š_éš±è—" not in pre_df.columns:
        pre_df["_é å¯©ç­‰åƒ¹ç´š_éš±è—"] = pre_df["é å¯©åˆ¤å®š"].apply(normalize_status_equiv)

    for _, prow in pre_df.iterrows():
        pid = str(prow.get("ç·¨è™Ÿ",""))
        pitem = str(prow.get("æª¢æ ¸é …ç›®",""))
        pori = str(prow.get("é å¯©åˆ¤å®š",""))
        peq = str(prow.get("_é å¯©ç­‰åƒ¹ç´š_éš±è—",""))
        matched = sys_idx.get(pid)
        matched_id = pid
        if not matched:
            best_id, best_ratio = fuzzy_match(list(sys_idx.keys()), pid or pitem)
            if best_ratio >= 0.85 and best_id in sys_idx:
                matched = sys_idx[best_id]; matched_id = best_id
        if matched:
            diff = "ä¸€è‡´" if (matched.get("ç¬¦åˆæƒ…å½¢","") == peq) else "ä¸ä¸€è‡´"
            rows_out.append({
                "é¡åˆ¥": matched.get("é¡åˆ¥",""),
                "ç·¨è™Ÿ": matched_id,
                "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰": matched.get("æª¢æ ¸é …ç›®",""),
                "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰": pori,
                "é å¯©ç­‰åƒ¹ç´š": peq,
                "ç³»çµ±æª¢æ ¸çµæœ": matched.get("ç¬¦åˆæƒ…å½¢",""),
                "å·®ç•°åˆ¤å®š": diff,
                "å·®ç•°èªªæ˜/å»ºè­°": matched.get("æ”¹å–„å»ºè­°",""),
                "å°æ‡‰é æ¬¡/å‚™è¨»": prow.get("å°æ‡‰é æ¬¡/å‚™è¨»",""),
            })
        else:
            rows_out.append({
                "é¡åˆ¥": "",
                "ç·¨è™Ÿ": pid or "ï¼ˆæœªè­˜åˆ¥ï¼‰",
                "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰": pitem,
                "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰": pori,
                "é å¯©ç­‰åƒ¹ç´š": peq or "æœªæåŠ",
                "ç³»çµ±æª¢æ ¸çµæœ": "ï¼ˆç„¡å°æ‡‰ï¼‰",
                "å·®ç•°åˆ¤å®š": "é å¯©å¤šå‡º",
                "å·®ç•°èªªæ˜/å»ºè­°": "æ­¤é å¯©é …ç›®æ–¼ç³»çµ±æª¢æ ¸æ¸…å–®ä¸­ç„¡ç›´æ¥å°æ‡‰ï¼Œè«‹äººå·¥ç¢ºèªã€‚",
                "å°æ‡‰é æ¬¡/å‚™è¨»": prow.get("å°æ‡‰é æ¬¡/å‚™è¨»",""),
            })

    pre_ids = set([str(x).strip() for x in pre_df.get("ç·¨è™Ÿ", pd.Series(dtype=str)).tolist() if str(x).strip()])
    for _, srow in sys_df.iterrows():
        sid = str(srow.get("ç·¨è™Ÿ",""))
        if sid and sid not in pre_ids:
            rows_out.append({
                "é¡åˆ¥": srow.get("é¡åˆ¥",""),
                "ç·¨è™Ÿ": srow.get("ç·¨è™Ÿ",""),
                "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰": srow.get("æª¢æ ¸é …ç›®",""),
                "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰": "",
                "é å¯©ç­‰åƒ¹ç´š": "æœªæåŠ",
                "ç³»çµ±æª¢æ ¸çµæœ": srow.get("ç¬¦åˆæƒ…å½¢",""),
                "å·®ç•°åˆ¤å®š": "ç³»çµ±å¤šå‡º",
                "å·®ç•°èªªæ˜/å»ºè­°": "é å¯©æœªæ¶µè“‹æ­¤é …ï¼Œå»ºè­°è£œåˆ—æˆ–æ–¼æœƒå¯©æ™‚æç¤ºæ‰¿è¾¦æ³¨æ„ã€‚",
                "å°æ‡‰é æ¬¡/å‚™è¨»": "",
            })

    out = pd.DataFrame(rows_out)
    try:
        out["ä¸»ç¢¼"]   = out["ç·¨è™Ÿ"].str.extract(r'^([A-F])')[0]
        out["å­ç¢¼ä¸»"] = pd.to_numeric(out["ç·¨è™Ÿ"].str.extract(r'^A-F')[0], errors="coerce")
        out["å­ç¢¼æ¬¡"] = pd.to_numeric(out["ç·¨è™Ÿ"].str.extract(r'^[A-F]\d+\.(\d+)')[0], errors="coerce")
        code_order = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5}
        out["ä¸»åº"]  = out["ä¸»ç¢¼"].map(code_order).fillna(9)
        out = out.sort_values(["ä¸»åº","å­ç¢¼ä¸»","å­ç¢¼æ¬¡","ç·¨è™Ÿ"], kind="mergesort").drop(columns=["ä¸»ç¢¼","å­ç¢¼ä¸»","å­ç¢¼æ¬¡","ä¸»åº"])
    except Exception:
        pass
    return out

# ==================== ä¸»ç¨‹å¼ ====================
def main():
    st.set_page_config(page_title="ğŸ“‘ è³‡è¨Šæœå‹™æ¡è³¼ RFP/å¥‘ç´„å¯©æŸ¥ç³»çµ±(LLMç‰ˆ)", layout="wide")
    st.title("ğŸ“‘ è³‡è¨Šæœå‹™æ¡è³¼ RFP/å¥‘ç´„å¯©æŸ¥ç³»çµ±(LLMç‰ˆ)")

    # RFP/å¥‘ç´„ PDFï¼ˆå¿…å¡«ï¼‰
    uploaded_files = st.file_uploader("ğŸ“¥ ä¸Šå‚³ RFP/å¥‘ç´„ PDFï¼ˆå¯è¤‡é¸ï¼‰", type=["pdf"], accept_multiple_files=True)
    # é å…ˆå¯©æŸ¥è¡¨ PDFï¼ˆå¯ç•¥éï¼‰
    pre_files = st.file_uploader("ğŸ“¥ ä¸Šå‚³ã€åŸ·è¡Œå–®ä½é å…ˆå¯©æŸ¥è¡¨ã€PDFï¼ˆå¯è¤‡é¸/å¯ç•¥éï¼‰", type=["pdf"], accept_multiple_files=True)

    project_name = st.text_input("æ¡ˆä»¶/å°ˆæ¡ˆåç¨±ï¼ˆå°‡ç”¨æ–¼æª”åï¼‰", value="æœªå‘½åæ¡ˆä»¶")
    work_summary = st.text_input("å·¥ä½œæ‘˜è¦ï¼ˆå¯é¸ï¼Œç”¨æ–¼ç¬¬1é»è£œè¿°ï¼‰", value="")

    

    st.caption("æª¢æ ¸æ¨¡å¼ï¼šä¸€æ¬¡æ€§å¯©æŸ¥")

    if st.button("ğŸš€ é–‹å§‹å¯©æŸ¥", disabled=not uploaded_files):
        checklist_all = build_rfp_checklist()
        progress_text = st.empty()
        progress_bar = st.progress(0)

        def set_progress(p, msg):
            progress_bar.progress(max(0, min(int(p), 100)))
            progress_text.write(msg)

        # 1) è§£æ RFP/å¥‘ç´„ PDF
        set_progress(5, "ğŸ“„ è§£æèˆ‡å½™æ•´ RFP/å¥‘ç´„ æ–‡ä»¶æ–‡å­—â€¦")
        corpora = []
        total_files = len(uploaded_files)
        st.info("ğŸ“„ é–‹å§‹è§£æ RFP/å¥‘ç´„ PDF æª”æ¡ˆâ€¦")
        for i, f in enumerate(uploaded_files):
            set_progress(5 + int(((i + 1)/max(1, total_files)) * 25), f"ğŸ“„ è§£æ {f.name} ({i+1}/{total_files})â€¦")
            pdf_bytes = f.read()
            text = extract_text_with_headers(pdf_bytes, f.name)
            if not text.strip():
                st.warning(f"âš ï¸ {f.name} å¯èƒ½æ˜¯æƒæå½±åƒ PDFï¼Œç„¡æ³•ç›´æ¥æŠ½æ–‡å­—ã€‚è«‹æä¾›å¯æœå°‹çš„ PDFã€‚")
            corpora.append(text)
        corpus_text = "\n\n".join(corpora)

        # 2) è§£æ é å…ˆå¯©æŸ¥è¡¨ï¼ˆå¯ç•¥éï¼›UIä¸é¡¯ç¤ºï¼Œåƒ…ä¾›å·®ç•°å°ç…§ï¼‰
        set_progress(32, "ğŸ§© è™•ç†é å…ˆå¯©æŸ¥è¡¨â€¦")
        pre_df = pd.DataFrame()
        if pre_files:
            pre_texts = []
            for pf in pre_files:
                if is_pdf(pf.name):
                    st.write(f"ğŸ“„ æ­£åœ¨è™•ç†ï¼š{pf.name}")
                    pbytes = pf.read()
                    ptext = extract_text_with_headers(pbytes, pf.name)
                    if ptext.strip():
                        pre_texts.append(ptext)
                    else:
                        st.warning(f"âš ï¸ {pf.name} å¯èƒ½æ˜¯æƒæå½±åƒ PDFï¼Œç„¡æ³•ç›´æ¥æŠ½æ–‡å­—ã€‚è«‹æä¾›å¯æœå°‹ PDFã€‚")
            if pre_texts:
                st.info("ğŸ“„ é–‹å§‹è§£æé å¯©è¡¨ PDF æª”æ¡ˆâ€¦")
                pre_corpus = "\n\n".join(pre_texts)
                prompt = make_precheck_parse_prompt(pre_corpus)
                try:
                    if model is None:
                        raise RuntimeError("LLM æœªå•Ÿç”¨ï¼ˆç¼ºå°‘ GOOGLE_API_KEYï¼‰")
                    st.info("ğŸ¤– å‘¼å«æ¨¡å‹é€²è¡Œé å¯©è¡¨çµæ§‹åŒ–è¾¨è­˜â€¦")
                    resp = model.generate_content(prompt)
                    rows = parse_precheck_json(resp.text)
                    if rows:
                        pre_df = precheck_rows_to_df(rows)
                except Exception as e:
                    st.warning(f"âš ï¸ é å¯©è¡¨è§£æå¤±æ•—ï¼š{e}")
            else:
                st.info("â„¹ï¸ æœªä¸Šå‚³æˆ–æœªæˆåŠŸè¾¨è­˜ä»»ä½•é å¯©è¡¨å…§å®¹ã€‚")
        else:
            st.info("â„¹ï¸ æœªä¸Šå‚³æˆ–æœªæˆåŠŸè¾¨è­˜ä»»ä½•é å¯©è¡¨å…§å®¹.")

        set_progress(35, "ğŸ§  æª¢æ ¸æº–å‚™ä¸­â€¦")

        # 3) å…ˆæŠ½å–é ç®— â†’ å†ç”Ÿæˆå»ºè­°å›è¦†å…§å®¹
        st.subheader("ğŸ“ å»ºè­°å›è¦†å…§å®¹ï¼ˆLLMè‡ªå‹•ç”Ÿæˆï¼‰")
        kb_context = kb_to_context(kb_items, selected_kb_ids)
        budget_wan, ev = llm_extract_budget(corpus_text)
        budget_input = st.text_input("é ç®—é‡‘é¡ï¼ˆè¬å…ƒï¼›å¯æ‰‹å‹•è¦†è“‹ï¼‰", value=budget_wan or "XXX")
        try:
            if model is None:
                raise RuntimeError("LLM æœªå•Ÿç”¨ï¼ˆç¼ºå°‘ GOOGLE_API_KEYï¼‰")
            prompt = make_reply_prompt(corpus_text, kb_context, budget_input, work_summary, max_points=6)
            resp = model.generate_content(prompt)
            reply_text = (resp.text or "").strip()
            st.text_area("å›è¦†å…§å®¹ï¼ˆLLMè¼¸å‡ºï¼‰", reply_text, height=300)
        except Exception as e:
            st.warning(f"LLM ç”¢ç”Ÿå¤±æ•—ï¼š{e}")

        # 4) ä¸€æ¬¡æ€§å¯©æŸ¥
        all_results: List[Dict[str, Any]] = []
        st.info("ğŸ§ª åŸ·è¡Œç³»çµ±æª¢æ ¸æ¨¡å¼ï¼šä¸€æ¬¡æ€§å¯©æŸ¥")
        groups = [("ABCDE", checklist_all)] if checklist_all else []
        total_batches = len(groups)
        for bi, (code, items) in enumerate(groups):
            set_progress(35 + int(((bi + 1)/max(1, total_batches)) * 55), f"ğŸ” ä¸€æ¬¡æ€§å¯©æŸ¥ï¼ˆ{code}ï¼‰â€¦ å…± {len(items)} é …")
            prompt = make_batch_prompt(code, items, corpus_text)
            try:
                if model is None:
                    raise RuntimeError("LLM æœªå•Ÿç”¨ï¼ˆç¼ºå°‘ GOOGLE_API_KEYï¼‰")
                resp = model.generate_content(prompt)
                arr = parse_json_array(resp.text)
            except Exception:
                arr = []
            allowed_ids = {it['id'] for it in items}
            id_to_meta = {it['id']: it for it in items}
            normalized = []
            for d in arr if isinstance(arr, list) else []:
                if not isinstance(d, dict):
                    continue
                rid = d.get('id')
                if rid not in allowed_ids:
                    continue
                meta = id_to_meta[rid]
                normalized.append({
                    'id': rid,
                    'category': d.get('category', meta['category']),
                    'item': d.get('item', meta['item']),
                    'compliance': d.get('compliance', ''),
                    'evidence': d.get('evidence', {}),
                    'recommendation': d.get('recommendation', ''),
                })
            returned_ids = {x['id'] for x in normalized}
            for it in items:
                if it['id'] not in returned_ids:
                    normalized.append({
                        'id': it['id'], 'category': it['category'], 'item': it['item'],
                        'compliance': 'æœªæåŠ', 'evidence': {}, 'recommendation': ''
                    })
            all_results.extend(normalized)

        # 5) æª¢æ ¸çµæœ â†’ DataFrameï¼ˆUIä¸é¡¯ç¤ºï¼‰
        set_progress(92, "ğŸ“¦ å½™æ•´èˆ‡è½‰è¡¨æ ¼â€¦")
        df = to_dataframe(all_results)
        st.success("âœ… å¯©æŸ¥å®Œæˆ")

        # 6) å·®ç•°å°ç…§ï¼ˆåƒ…é¡¯ç¤ºä¸ä¸€è‡´/ç¼ºæ¼ï¼‰
        cmp_df = pd.DataFrame()
        if not pre_df.empty and not df.empty:
            st.info("ğŸ“‹ å»ºç«‹é å¯©èˆ‡ç³»çµ±æª¢æ ¸çš„å·®ç•°å°ç…§è¡¨â€¦")
            cmp_df = build_compare_table(sys_df=df, pre_df=pre_df)
        st.subheader("ğŸ§® å·®ç•°å°ç…§è¡¨ï¼ˆåªé¡¯ç¤ºä¸ä¸€è‡´/ç¼ºæ¼ï¼‰")
        view_df = cmp_df[cmp_df["å·®ç•°åˆ¤å®š"] != "ä¸€è‡´"] if not cmp_df.empty else pd.DataFrame(columns=["é¡åˆ¥", "ç·¨è™Ÿ", "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰", "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰", "å°æ‡‰é æ¬¡/å‚™è¨»", "ç³»çµ±æª¢æ ¸çµæœ", "å·®ç•°èªªæ˜/å»ºè­°"])        
        cmp_display_cols = ["é¡åˆ¥", "ç·¨è™Ÿ", "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰", "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰", "å°æ‡‰é æ¬¡/å‚™è¨»", "ç³»çµ±æª¢æ ¸çµæœ", "å·®ç•°èªªæ˜/å»ºè­°"]
        view_df = view_df[cmp_display_cols] if not view_df.empty else view_df
        search_term = st.text_input("ğŸ” æœå°‹æª¢æ ¸é …ç›®")
        if not view_df.empty and search_term:
            view_df = view_df[view_df["æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰"].str.contains(search_term, case=False, na=False)]
        st.data_editor(
            view_df,
            use_container_width=True,
            hide_index=True,
            disabled=["é¡åˆ¥", "ç·¨è™Ÿ", "æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰", "ç³»çµ±æª¢æ ¸çµæœ"],
            column_config={
                "é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰": st.column_config.SelectboxColumn(
                    "é å¯©åˆ¤å®š", options=["ç¬¦åˆ", "ä¸é©ç”¨", ""], required=False)
            }
        )

        # 7) Excel åŒ¯å‡ºï¼ˆåƒ…å·®ç•°å°ç…§ï¼‰
        try:
            from openpyxl.styles import Alignment
            xbio = io.BytesIO()
            st.info("ğŸ“ åŒ¯å‡º Excelï¼ˆåƒ…å·®ç•°å°ç…§ï¼‰â€¦")
            with pd.ExcelWriter(xbio, engine='openpyxl') as writer:
                if 'cmp_df' in locals() and not cmp_df.empty:
                    cmp_df.to_excel(writer, index=False, sheet_name='å·®ç•°å°ç…§')
                    ws3 = writer.sheets['å·®ç•°å°ç…§']
                    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    empty_df = pd.DataFrame(columns=["é¡åˆ¥","ç·¨è™Ÿ","æª¢æ ¸é …ç›®ï¼ˆç³»çµ±åŸºæº–ï¼‰","é å¯©åˆ¤å®šï¼ˆåŸå­—ï¼‰","å°æ‡‰é æ¬¡/å‚™è¨»","ç³»çµ±æª¢æ ¸çµæœ","å·®ç•°èªªæ˜/å»ºè­°","å·®ç•°åˆ¤å®š"])
                    empty_df.to_excel(writer, index=False, sheet_name='å·®ç•°å°ç…§')
            xbio.seek(0)
            st.download_button(
                label="ğŸ“¥ ä¸‹è¼‰ Excelï¼ˆå·®ç•°å°ç…§ï¼‰",
                data=xbio.getvalue(),
                file_name=f"{project_name}_RFP_Contract_Compare.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        except Exception as e:
            st.warning(f"Excel åŒ¯å‡ºå¤±æ•—ï¼š{e}")

        progress_text.empty(); progress_bar.empty()


if __name__ == '__main__':
    main()
